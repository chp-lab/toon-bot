from flask.helpers import send_file
from requests.api import head
from hooking import Hooking
from flask_restful import Resource
from flask import request
import requests
from database import Database
from module import Module
from datetime import datetime
import urllib3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
from database import Database
from flask import stream_with_context, Response
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class Export_excel(Resource):
    TAG = "ExportAsExcel:"
    file_path = "/home/user01/ibeacon-dev/tmp"
    # file_path = "./ibeacon-dev/tmp"
    # file_path = "/tmp"
    tmp_file_name = ""
    new_file_name = "default"

    def __init__(self):
        # clear file extension
        # file_name = args.get("file_name")
        self.file_path = self.file_path
        print(self.TAG, self.file_path, " file_path")
        if (not os.path.exists(self.file_path)):
            os.mkdir(self.file_path)
            # print(self.TAG, file_path, "dir created")
        # self.new_file_name = file_name.replace(".xlsx", "")
        filename = request.args['file_name']
        if((filename == None) || (filename == "undefined")):
            database = Database()
            cmd = """SELECT CURRENT_TIMESTAMP as file_name FROM users"""
            res = database.getData(cmd)
            my_file_name = res[0]['result'][0]['file_name']
            filename = "report_" + my_file_name
        self.new_file_name = filename + ".xlsx"
        # create full path
        self.tmp_file_name = self.file_path + "/" + self.new_file_name
        print(self.tmp_file_name)

    # user who open file location

    def open_file_location(self):
        # print(self.TAG, "%s$Location:" %(user), self.tmp_file_name)
        return {"file_path": self.file_path, "file_name": self.new_file_name}

    # def read_file(self):
    #     print(self.TAG, "read file", self.tmp_file_name)
    def export_file(self, table):
        # print(self.TAG, "exportFile")
        wb = Workbook()
        # ws = wb.create_sheet(0)
        ws = wb.active
        # ws.title = "Sheet"
        # print(self.TAG, "Exporting to", self.tmp_file_name)
        for col in table["head"]:
            print(col)
        try:
            print(self.TAG, "heads=", table["head"])
            if (table["head"] is None):
                return False
            heads = ["No."] + table["head"]
            #             print(self.TAG, "heads=",heads)
            if ("reverse_cut" in table):
                print("reverse cut found")
                heads.pop()
            ws.append(heads)
            # for data in table["data"]:
            #     ws.append(data)
            for col_name in table:
                print(self.TAG, col_name)
            if (table["data"] is not None):
                data_export = table["data"]
                print(self.TAG, "data type=", type(table["data"]))
                if (isinstance(data_export, str)):
                    data_export = json.loads(data_export)
                cut_num = 0
                if ("cut" in table):
                    print("cut=", table["cut"])
                    num_col = 0
                    try:
                        cut_num = int(table["cut"])
                        if (len(data_export) > 0):
                            num_col = len(data_export[0])
                        if("reverse_cut" in table):
                            num_col = num_col - int(table["reverse_cut"])
                    except Exception as err2:
                        print(self.TAG, "DB0X01 error on cutting, err=", err2)
                        return False

                    for i in range(len(data_export)):
                        tmp_data = [i + 1] + data_export[i][cut_num:num_col]
                        ws.append(tmp_data)
                else:
                    for i in range(len(data_export)):
                        tmp_data = [i + 1] + data_export[i]
                        ws.append(tmp_data)

                if (("sum_at_table_footer_lbl" in table) and (len(data_export) > 0)):
                    num_col = len(data_export[0]) + 1 - cut_num
                    # add no col
                    if(num_col > 2):
                        blank_footer = []
                        for i in range(num_col - 2):
                            blank_footer.append("")
                    blank_footer.append(table["sum_at_table_footer_lbl"])
                    tmp_order_number = []
                    # fixed index
                    sum = 0
                    order_number_id = 8
                    status_id = 9
                    price_vat_id = 21
                    try:
                        order_number_id = order_number_id + cut_num - 1
                        status_id = status_id + cut_num - 1
                        price_vat_id = price_vat_id + cut_num - 1

                        for row in data_export:
                            # print("order_number=", row[order_number_id])
                            # print("status=", row[status_id])
                            # print("price=", row[price_vat_id])
                            if(row[order_number_id] not in tmp_order_number):
                                tmp_order_number.append(row[order_number_id])
                                # print("new order found!")
                                if(row[status_id] == "approved"):
                                    # print("approved")
                                    sum = sum + float(row[price_vat_id])
                    except Exception as err2:
                        print(self.TAG, "DB0X02 err=", err2)
                        # return False

                    print("sum=", sum)

                    blank_footer.append(sum)
                    ws.append(blank_footer)
            #                 for i in range(len(heads)):
            #                     ws.cell(1, i + 1).font = Font(bold=True)
            #                     ws.column_dimensions[chr(ord('A') + i)].width = 25
            # ws.column_dimensions[get_column_letter(i + 1)].width = 25

            # edit sheet style
            try:
                for i in range(len(heads)):
                    ws.cell(1, i + 1).font = Font(bold=True)
                    ws.column_dimensions[get_column_letter(i + 1)].width = 25
                    # get last row
                    ws.cell(len(data_export) + 2, i + 1).font = Font(bold=True)
            except Exception as err2:
                print(self.TAG, "DB0X03 err=", err2)
                return False

            wb.save(self.tmp_file_name)
        except Exception as err:
            print(self.TAG, "error on exporting excel")
            print(self.TAG, "DB0X04 err=", err)
            return False

        return True
    # def delete_file(self):
    #     print(self.TAG, "delete file", self.tmp_file_name)
    #     try:
    #         os.remove("demofile.txt")
    #     except Exception as err:
    #         print(self.TAG, "err=", err)

    # def edit_file(self, table):
    #     print(self.TAG, "edit file", self.tmp_file_name)

    def download_file(self, file_path, file_name):
        # self.file_path = file_path
        # print("DOWNLOAD ", self.file_path, " Thissssssssssssssss file pathhhh")
        # self.new_file_name = file_name + ".xlsx"
        # print("DOWNLOAD ", self.new_file_name,
        #       " Thissssssssssssssss new_file_name")
        # self.tmp_file_name = self.file_path + "/" + self.new_file_name
        # print("DOWNLOAD ", self.tmp_file_name,
        #       " Thissssssssssssssss tmp_file_name")
        if os.path.exists(file_path):
            list_filename = os.listdir(file_path)
            new_filename = file_name + ".xlsx"
            print("DOWNLOAD ", list_filename, "list file_name")
            print()
            if new_filename in list_filename:
                path_name_file = file_path + "/" + new_filename
                print("DOWNLOAD ", path_name_file, "file_name and path_name")
                return send_file(path_name_file, as_attachment=True)
            else:
                return "file name not found"
        else:
            return "path name not found"

    def post(self):
        # args = request.args
        # excel_file = Export_excel(args)
        # excel_file = Export_excel(self.new_file_name)
        # excel_file = self.__init__(self, args)
        # print(request.json)
        # print(request.json['head'])
        # print(request.json['data'])

        dataFile = {
            "head": request.json['head'],
            "data": request.json['data']
        }
        excel_file = self.export_file(dataFile)
        file_location = self.open_file_location()
        print("Location: ", file_location)

        return file_location

    def get(self):
        filename = request.args['file_name']
        filepath = request.args['file_path']
        if filename is not None and filepath is not None:
            excel_file = self.download_file(filepath, filename)
            return excel_file
        # else:
        #     return "Request Param Not found"
        # return Response(stream_with_context(self.download_file(filename, filepath)))
